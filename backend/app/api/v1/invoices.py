import io
import json
from datetime import date

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import and_, or_, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_user
from app.core.config import settings
from app.db.session import get_db
from app.models.building import Building
from app.models.invoice import Invoice
from app.models.price_config import PriceConfig
from app.models.reading import MeterReading
from app.models.room import Room
from app.models.user import User
from app.schemas.invoice import (
    InvoiceGenerateRequest,
    InvoiceGenerateResponse,
    InvoiceGenerateRoomResult,
    InvoiceResponse,
)
from app.services.billing_service import calculate_invoice
from app.services.pdf_service import generate_invoice_pdf

router = APIRouter(prefix="/invoices", tags=["Invoices"])


def _month_bounds(invoice_month: str) -> tuple[date, date]:
    try:
        month_start = date.fromisoformat(f"{invoice_month}-01")
    except ValueError as exc:
        raise HTTPException(
            status_code=422,
            detail="Kỳ hóa đơn phải theo định dạng YYYY-MM",
        ) from exc

    if len(invoice_month) != 7 or month_start.strftime("%Y-%m") != invoice_month:
        raise HTTPException(
            status_code=422,
            detail="Kỳ hóa đơn phải theo định dạng YYYY-MM",
        )

    if month_start.month == 12:
        next_month = date(month_start.year + 1, 1, 1)
    else:
        next_month = date(month_start.year, month_start.month + 1, 1)
    return month_start, next_month


async def _get_owned_building(
    db: AsyncSession, building_id: int, owner_id: int
) -> Building:
    result = await db.execute(
        select(Building).where(
            Building.id == building_id,
            Building.owner_id == owner_id,
        )
    )
    building = result.scalar_one_or_none()
    if not building:
        raise HTTPException(status_code=404, detail="Tòa nhà không tồn tại")
    return building


def _invoice_response(invoice: Invoice, room: Room) -> InvoiceResponse:
    response = InvoiceResponse.model_validate(invoice)
    response.room_number = room.room_number
    response.resident_name = room.resident_name
    return response


def _excel_safe(value: str) -> str:
    """Force user-controlled strings to remain text when opened in a spreadsheet."""
    if value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


@router.post("/generate", response_model=InvoiceGenerateResponse)
async def generate_invoices(
    body: InvoiceGenerateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    month_start, next_month = _month_bounds(body.invoice_month)
    await _get_owned_building(db, body.building_id, current_user.id)

    # Get price config
    config_result = await db.execute(
        select(PriceConfig).where(
            PriceConfig.id == body.price_config_id,
            PriceConfig.is_active == True,
        )
    )
    price_config = config_result.scalar_one_or_none()
    if not price_config:
        raise HTTPException(status_code=404, detail="Bảng giá không tồn tại")

    # Get all active rooms in building
    rooms_result = await db.execute(
        select(Room).where(Room.building_id == body.building_id, Room.is_active == True)
    )
    rooms = rooms_result.scalars().all()

    if not rooms:
        raise HTTPException(status_code=400, detail="Không có phòng nào trong tòa nhà")

    invoices_created = []
    room_results: list[InvoiceGenerateRoomResult] = []
    total_amount = 0.0

    for room in rooms:
        # Check if invoice already exists
        existing = await db.execute(
            select(Invoice).where(
                Invoice.room_id == room.id,
                Invoice.invoice_month == body.invoice_month,
            )
        )
        existing_invoice = existing.scalar_one_or_none()
        if existing_invoice:
            room_results.append(
                InvoiceGenerateRoomResult(
                    room_id=room.id,
                    room_number=room.room_number,
                    status="skipped",
                    invoice_id=existing_invoice.id,
                    detail="Hóa đơn tháng này đã tồn tại",
                )
            )
            continue

        # Get latest approved reading for this month
        current_result = await db.execute(
            select(MeterReading)
            .where(
                MeterReading.room_id == room.id,
                MeterReading.status == "approved",
                MeterReading.reading_date >= month_start,
                MeterReading.reading_date < next_month,
            )
            .order_by(MeterReading.reading_date.desc(), MeterReading.id.desc())
            .limit(1)
        )
        current_reading_obj = current_result.scalar_one_or_none()
        if not current_reading_obj:
            room_results.append(
                InvoiceGenerateRoomResult(
                    room_id=room.id,
                    room_number=room.room_number,
                    status="skipped",
                    detail="Chưa có chỉ số đã duyệt trong tháng",
                )
            )
            continue

        # Get previous reading
        prev_result = await db.execute(
            select(MeterReading)
            .where(
                MeterReading.room_id == room.id,
                MeterReading.status == "approved",
                or_(
                    MeterReading.reading_date < current_reading_obj.reading_date,
                    and_(
                        MeterReading.reading_date == current_reading_obj.reading_date,
                        MeterReading.id < current_reading_obj.id,
                    ),
                ),
            )
            .order_by(MeterReading.reading_date.desc(), MeterReading.id.desc())
            .limit(1)
        )
        prev_reading_obj = prev_result.scalar_one_or_none()
        previous_value = prev_reading_obj.meter_value if prev_reading_obj else room.initial_reading

        current_value = current_reading_obj.meter_value
        consumption = current_value - previous_value

        if consumption < 0:
            room_results.append(
                InvoiceGenerateRoomResult(
                    room_id=room.id,
                    room_number=room.room_number,
                    status="error",
                    detail="Chỉ số mới nhỏ hơn chỉ số cũ",
                )
            )
            continue

        # Calculate
        try:
            calc_result = calculate_invoice(
                consumption=consumption,
                pricing_type=price_config.pricing_type,
                config_json=price_config.config_json,
                additional_fees=body.additional_fees or None,
            )
        except (KeyError, TypeError, ValueError) as exc:
            room_results.append(
                InvoiceGenerateRoomResult(
                    room_id=room.id,
                    room_number=room.room_number,
                    status="error",
                    detail=f"Không tính được tiền điện: {exc}",
                )
            )
            continue

        invoice = Invoice(
            room_id=room.id,
            reading_id=current_reading_obj.id,
            invoice_month=body.invoice_month,
            previous_reading=previous_value,
            current_reading=current_value,
            consumption=consumption,
            price_breakdown=json.dumps(calc_result["price_breakdown"], ensure_ascii=False),
            electricity_amount=calc_result["electricity_amount"],
            additional_fees=json.dumps(calc_result["additional_fees"], ensure_ascii=False) if calc_result["additional_fees"] else None,
            total_amount=calc_result["total_amount"],
        )
        try:
            async with db.begin_nested():
                db.add(invoice)
                await db.flush()
        except IntegrityError:
            existing = await db.execute(
                select(Invoice).where(
                    Invoice.room_id == room.id,
                    Invoice.invoice_month == body.invoice_month,
                )
            )
            existing_invoice = existing.scalar_one_or_none()
            if existing_invoice:
                room_results.append(
                    InvoiceGenerateRoomResult(
                        room_id=room.id,
                        room_number=room.room_number,
                        status="skipped",
                        invoice_id=existing_invoice.id,
                        detail="Hóa đơn tháng này đã được tạo đồng thời",
                    )
                )
                continue
            raise

        resp = _invoice_response(invoice, room)
        invoices_created.append(resp)
        total_amount += calc_result["total_amount"]
        room_results.append(
            InvoiceGenerateRoomResult(
                room_id=room.id,
                room_number=room.room_number,
                status="created",
                invoice_id=invoice.id,
                detail="Đã tạo hóa đơn",
            )
        )

    await db.commit()

    return InvoiceGenerateResponse(
        total_invoices=len(invoices_created),
        total_amount=total_amount,
        invoices=invoices_created,
        total_skipped=sum(result.status == "skipped" for result in room_results),
        total_errors=sum(result.status == "error" for result in room_results),
        results=room_results,
    )


@router.get("", response_model=list[InvoiceResponse])
async def list_invoices(
    building_id: int | None = None,
    invoice_month: str | None = None,
    sent_status: str | None = None,
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=200, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = (
        select(Invoice, Room)
        .join(Room, Invoice.room_id == Room.id)
        .join(Building, Room.building_id == Building.id)
        .where(Building.owner_id == current_user.id)
        .order_by(Invoice.created_at.desc())
    )

    if building_id is not None:
        await _get_owned_building(db, building_id, current_user.id)
        query = query.where(Room.building_id == building_id)

    if invoice_month:
        _month_bounds(invoice_month)
        query = query.where(Invoice.invoice_month == invoice_month)
    if sent_status:
        query = query.where(Invoice.sent_status == sent_status)

    result = await db.execute(query.offset(offset).limit(limit))
    return [_invoice_response(invoice, room) for invoice, room in result.all()]


@router.get("/{invoice_id}", response_model=InvoiceResponse)
async def get_invoice(
    invoice_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(
        select(Invoice, Room)
        .join(Room, Invoice.room_id == Room.id)
        .join(Building, Room.building_id == Building.id)
        .where(
            Invoice.id == invoice_id,
            Building.owner_id == current_user.id,
        )
    )
    row = result.one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Hóa đơn không tồn tại")
    invoice, room = row
    return _invoice_response(invoice, room)


@router.get("/{invoice_id}/pdf")
async def download_invoice_pdf(
    invoice_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(
        select(Invoice, Room)
        .join(Room, Invoice.room_id == Room.id)
        .join(Building, Room.building_id == Building.id)
        .where(
            Invoice.id == invoice_id,
            Building.owner_id == current_user.id,
        )
    )
    row = result.one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Hóa đơn không tồn tại")
    invoice, room = row

    pdf_bytes = generate_invoice_pdf(
        invoice_id=invoice.id,
        invoice_month=invoice.invoice_month,
        room_number=room.room_number,
        resident_name=room.resident_name,
        previous_reading=invoice.previous_reading,
        current_reading=invoice.current_reading,
        consumption=invoice.consumption,
        price_breakdown_json=invoice.price_breakdown,
        electricity_amount=invoice.electricity_amount,
        additional_fees_json=invoice.additional_fees,
        total_amount=invoice.total_amount,
        management_unit=settings.PAYMENT_MANAGEMENT_UNIT,
        bank_account=settings.PAYMENT_BANK_ACCOUNT,
        bank_name=settings.PAYMENT_BANK_NAME,
        account_holder=settings.PAYMENT_ACCOUNT_HOLDER,
    )

    month_safe = invoice.invoice_month.replace("-", "")
    room_safe = room.room_number.replace(" ", "_").replace("/", "-")
    filename = f"hoa-don-{room_safe}-{month_safe}.pdf"

    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@router.get("/export/excel")
async def export_excel(
    invoice_month: str,
    building_id: int | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    _month_bounds(invoice_month)

    # Build query — filter by specific building or all buildings owned by user
    q = (
        select(Invoice, Room, Building)
        .join(Room, Invoice.room_id == Room.id)
        .join(Building, Room.building_id == Building.id)
        .where(
            Building.owner_id == current_user.id,
            Invoice.invoice_month == invoice_month,
        )
    )
    if building_id is not None:
        q = q.where(Building.id == building_id)
    q = q.order_by(Building.name, Room.room_number)

    result = await db.execute(q)
    rows = result.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Tháng {invoice_month}"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1E40AF")
    total_font = Font(bold=True, size=11)
    total_fill = PatternFill("solid", fgColor="DBEAFE")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title row
    title = f"BÁO CÁO TIỀN ĐIỆN THÁNG {invoice_month}"
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(bold=True, size=14, color="1E3A5F")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.append([])  # blank row

    # Headers
    headers = [
        "STT", "Tòa nhà", "Phòng", "Cư dân",
        "Chỉ số cũ", "Chỉ số mới", "Tiêu thụ (kWh)",
        "Đơn giá (đ/kWh)", "Tiền điện (đ)", "Phụ phí (đ)",
        "Tổng cộng (đ)",
    ]
    ws.append(headers)
    header_row = ws.max_row
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[header_row].height = 22

    # Data rows
    total_consumption = 0
    total_electricity = 0.0
    total_fees = 0.0
    total_amount = 0.0

    for i, (inv, room, building) in enumerate(rows, 1):
        fees = 0.0
        if inv.additional_fees:
            fees_data = json.loads(inv.additional_fees)
            fees = float(sum(fees_data.values()))

        # Derive average unit price from price_breakdown if available
        unit_price = 0.0
        if inv.price_breakdown:
            try:
                bd = json.loads(inv.price_breakdown)
                if isinstance(bd, list) and bd:
                    unit_price = sum(t.get("amount", 0) for t in bd) / inv.consumption if inv.consumption else 0
                elif isinstance(bd, dict):
                    unit_price = bd.get("unit_price", 0) or (inv.electricity_amount / inv.consumption if inv.consumption else 0)
            except Exception:
                pass
        if unit_price == 0 and inv.consumption:
            unit_price = round(inv.electricity_amount / inv.consumption, 2)

        data_row = [
            i,
            _excel_safe(building.name),
            _excel_safe(room.room_number),
            _excel_safe(room.resident_name or ""),
            inv.previous_reading,
            inv.current_reading,
            inv.consumption,
            round(unit_price),
            inv.electricity_amount,
            fees,
            inv.total_amount,
        ]
        ws.append(data_row)
        row_idx = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            if col_idx in (1, 5, 6, 7):
                cell.alignment = center
            if col_idx >= 5:
                cell.number_format = "#,##0"

        total_consumption += inv.consumption
        total_electricity += inv.electricity_amount
        total_fees += fees
        total_amount += inv.total_amount

    # Totals row
    ws.append([])
    ws.append([
        "", "TỔNG CỘNG", "", "",
        "", "", total_consumption,
        "", total_electricity,
        total_fees, total_amount,
    ])
    total_row = ws.max_row
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = border
        if col_idx >= 5:
            cell.number_format = "#,##0"

    # Column widths
    col_widths = [5, 18, 10, 20, 11, 11, 15, 16, 15, 13, 16]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"bao-cao-tien-dien-{invoice_month}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )

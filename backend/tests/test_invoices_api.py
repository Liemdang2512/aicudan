"""API safety-net tests for invoice period selection and export."""

from datetime import date
from io import BytesIO

import openpyxl
import pytest
from httpx import AsyncClient
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.invoice import Invoice
from app.models.reading import MeterReading
from app.models.room import Room


async def _add_reading(
    db_session: AsyncSession,
    room: Room,
    reading_date: date,
    meter_value: int,
    status: str = "approved",
) -> MeterReading:
    reading = MeterReading(
        room_id=room.id,
        reading_date=reading_date,
        meter_value=meter_value,
        status=status,
        confidence_score=1.0,
    )
    db_session.add(reading)
    await db_session.flush()
    return reading


@pytest.mark.asyncio
async def test_invoice_requires_an_approved_reading(
    client: AsyncClient,
    auth_headers: dict[str, str],
    db_session: AsyncSession,
    test_room: Room,
    test_fixed_price_config,
):
    await _add_reading(db_session, test_room, date(2025, 1, 31), 1100, "needs_review")
    await db_session.commit()

    response = await client.post(
        "/api/v1/invoices/generate",
        headers=auth_headers,
        json={
            "building_id": test_room.building_id,
            "invoice_month": "2025-01",
            "price_config_id": test_fixed_price_config.id,
        },
    )

    assert response.status_code == 200
    assert response.json()["total_invoices"] == 0
    assert response.json()["invoices"] == []


@pytest.mark.asyncio
async def test_invoice_uses_latest_approved_reading_in_selected_month(
    client: AsyncClient,
    auth_headers: dict[str, str],
    db_session: AsyncSession,
    test_room: Room,
    test_fixed_price_config,
):
    previous = await _add_reading(db_session, test_room, date(2025, 1, 10), 1000)
    current = await _add_reading(db_session, test_room, date(2025, 1, 25), 1150)
    await _add_reading(db_session, test_room, date(2025, 2, 5), 1300)
    await db_session.commit()

    response = await client.post(
        "/api/v1/invoices/generate",
        headers=auth_headers,
        json={
            "building_id": test_room.building_id,
            "invoice_month": "2025-01",
            "price_config_id": test_fixed_price_config.id,
        },
    )

    assert response.status_code == 200
    assert response.json()["total_invoices"] == 1
    invoice = response.json()["invoices"][0]
    assert invoice["current_reading"] == current.meter_value
    assert invoice["previous_reading"] == previous.meter_value
    assert invoice["consumption"] == 150


@pytest.mark.asyncio
async def test_invoice_uses_previous_reading_by_date_then_id_on_same_day(
    client: AsyncClient,
    auth_headers: dict[str, str],
    db_session: AsyncSession,
    test_room: Room,
    test_fixed_price_config,
):
    previous = await _add_reading(db_session, test_room, date(2025, 1, 31), 1100)
    current = await _add_reading(db_session, test_room, date(2025, 1, 31), 1150)
    assert previous.id < current.id
    await db_session.commit()

    response = await client.post(
        "/api/v1/invoices/generate",
        headers=auth_headers,
        json={
            "building_id": test_room.building_id,
            "invoice_month": "2025-01",
            "price_config_id": test_fixed_price_config.id,
        },
    )

    assert response.status_code == 200
    invoice = response.json()["invoices"][0]
    assert invoice["previous_reading"] == 1100
    assert invoice["current_reading"] == 1150
    assert invoice["consumption"] == 50


@pytest.mark.asyncio
async def test_future_reading_cannot_create_invoice_for_selected_month(
    client: AsyncClient,
    auth_headers: dict[str, str],
    db_session: AsyncSession,
    test_room: Room,
    test_fixed_price_config,
):
    await _add_reading(db_session, test_room, date(2025, 2, 1), 1200)
    await db_session.commit()

    response = await client.post(
        "/api/v1/invoices/generate",
        headers=auth_headers,
        json={
            "building_id": test_room.building_id,
            "invoice_month": "2025-01",
            "price_config_id": test_fixed_price_config.id,
        },
    )

    assert response.status_code == 200
    assert response.json()["total_invoices"] == 0
    assert response.json()["total_skipped"] == 1
    assert response.json()["results"][0]["status"] == "skipped"


@pytest.mark.asyncio
async def test_invoice_falls_back_to_room_initial_reading(
    client: AsyncClient,
    auth_headers: dict[str, str],
    db_session: AsyncSession,
    test_room: Room,
    test_fixed_price_config,
):
    await _add_reading(db_session, test_room, date(2025, 1, 31), 1125)
    await db_session.commit()

    response = await client.post(
        "/api/v1/invoices/generate",
        headers=auth_headers,
        json={
            "building_id": test_room.building_id,
            "invoice_month": "2025-01",
            "price_config_id": test_fixed_price_config.id,
        },
    )

    assert response.status_code == 200
    invoice = response.json()["invoices"][0]
    assert invoice["previous_reading"] == test_room.initial_reading
    assert invoice["consumption"] == 125


@pytest.mark.asyncio
async def test_invoice_generation_is_idempotent_per_room_and_month(
    client: AsyncClient,
    auth_headers: dict[str, str],
    db_session: AsyncSession,
    test_room: Room,
    test_fixed_price_config,
):
    await _add_reading(db_session, test_room, date(2025, 1, 31), 1125)
    await db_session.commit()
    payload = {
        "building_id": test_room.building_id,
        "invoice_month": "2025-01",
        "price_config_id": test_fixed_price_config.id,
    }

    first = await client.post(
        "/api/v1/invoices/generate", headers=auth_headers, json=payload
    )
    second = await client.post(
        "/api/v1/invoices/generate", headers=auth_headers, json=payload
    )

    assert first.status_code == 200
    assert first.json()["total_invoices"] == 1
    assert first.json()["results"][0]["status"] == "created"
    assert second.status_code == 200
    assert second.json()["total_invoices"] == 0
    assert second.json()["total_skipped"] == 1
    assert second.json()["results"][0]["status"] == "skipped"
    assert second.json()["results"][0]["invoice_id"] == first.json()["invoices"][0]["id"]


@pytest.mark.asyncio
async def test_database_rejects_duplicate_invoice_for_room_and_month(
    db_session: AsyncSession,
    test_invoice: Invoice,
):
    duplicate = Invoice(
        room_id=test_invoice.room_id,
        reading_id=test_invoice.reading_id,
        invoice_month=test_invoice.invoice_month,
        previous_reading=test_invoice.previous_reading,
        current_reading=test_invoice.current_reading,
        consumption=test_invoice.consumption,
        electricity_amount=test_invoice.electricity_amount,
        total_amount=test_invoice.total_amount,
    )
    db_session.add(duplicate)

    with pytest.raises(IntegrityError):
        await db_session.commit()
    await db_session.rollback()


@pytest.mark.asyncio
@pytest.mark.parametrize("invoice_month", ["2025-13", "2025-1", "not-a-month"])
async def test_invoice_generation_rejects_invalid_month(
    client: AsyncClient,
    auth_headers: dict[str, str],
    test_room: Room,
    test_fixed_price_config,
    invoice_month: str,
):
    response = await client.post(
        "/api/v1/invoices/generate",
        headers=auth_headers,
        json={
            "building_id": test_room.building_id,
            "invoice_month": invoice_month,
            "price_config_id": test_fixed_price_config.id,
        },
    )

    assert response.status_code == 422


@pytest.mark.asyncio
async def test_invoice_endpoints_are_scoped_to_owner(
    client: AsyncClient,
    auth_headers: dict[str, str],
    test_invoice,
    second_invoice,
    second_building,
    test_fixed_price_config,
):
    list_response = await client.get("/api/v1/invoices", headers=auth_headers)
    export_response = await client.get(
        "/api/v1/invoices/export/excel",
        headers=auth_headers,
        params={"building_id": second_building.id, "invoice_month": "2025-01"},
    )
    generate_response = await client.post(
        "/api/v1/invoices/generate",
        headers=auth_headers,
        json={
            "building_id": second_building.id,
            "invoice_month": "2025-01",
            "price_config_id": test_fixed_price_config.id,
        },
    )

    assert list_response.status_code == 200
    assert [invoice["id"] for invoice in list_response.json()] == [test_invoice.id]
    assert second_invoice.id not in [invoice["id"] for invoice in list_response.json()]
    assert export_response.status_code == 404
    assert generate_response.status_code == 404


@pytest.mark.asyncio
async def test_export_endpoint_returns_only_requested_building_and_month(
    client: AsyncClient,
    auth_headers: dict[str, str],
    test_invoice,
    test_room: Room,
):
    response = await client.get(
        "/api/v1/invoices/export/excel",
        headers=auth_headers,
        params={"building_id": test_room.building_id, "invoice_month": "2025-01"},
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = openpyxl.load_workbook(BytesIO(response.content), data_only=True)
    rows = list(workbook.active.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1
    assert rows[0][1] == "101"


@pytest.mark.asyncio
@pytest.mark.parametrize("dangerous_prefix", ["=", "+", "-", "@"])
async def test_export_neutralizes_formula_injection_in_user_strings(
    client: AsyncClient,
    auth_headers: dict[str, str],
    db_session: AsyncSession,
    test_invoice,
    test_room: Room,
    dangerous_prefix: str,
):
    test_room.room_number = f"{dangerous_prefix}SUM(1,1)"
    test_room.resident_name = f"{dangerous_prefix}cmd"
    await db_session.commit()

    response = await client.get(
        "/api/v1/invoices/export/excel",
        headers=auth_headers,
        params={"building_id": test_room.building_id, "invoice_month": "2025-01"},
    )

    assert response.status_code == 200
    workbook = openpyxl.load_workbook(BytesIO(response.content), data_only=False)
    row = next(workbook.active.iter_rows(min_row=2, values_only=True))
    assert row[1] == f"'{dangerous_prefix}SUM(1,1)"
    assert row[2] == f"'{dangerous_prefix}cmd"
    assert workbook.active["B2"].data_type == "s"
    assert workbook.active["C2"].data_type == "s"


@pytest.mark.asyncio
async def test_invoice_list_exposes_bounded_pagination(
    client: AsyncClient,
    auth_headers: dict[str, str],
    test_invoice,
):
    response = await client.get(
        "/api/v1/invoices",
        headers=auth_headers,
        params={"offset": 0, "limit": 1},
    )
    invalid = await client.get(
        "/api/v1/invoices",
        headers=auth_headers,
        params={"limit": 0},
    )

    assert response.status_code == 200
    assert len(response.json()) == 1
    assert invalid.status_code == 422


@pytest.mark.asyncio
async def test_invoice_list_serializes_sending_status(
    client: AsyncClient,
    auth_headers: dict[str, str],
    db_session: AsyncSession,
    test_invoice: Invoice,
):
    test_invoice.sent_status = "sending"
    await db_session.commit()

    response = await client.get("/api/v1/invoices", headers=auth_headers)

    assert response.status_code == 200
    assert response.json()[0]["sent_status"] == "sending"
